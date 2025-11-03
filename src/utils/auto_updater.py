"""
Auto-Updater Service - Automatic device configuration synchronization
Controlla e sincronizza automaticamente la configurazione dei dispositivi all'avvio
"""

import logging
from pathlib import Path
from typing import Dict, Optional
from datetime import datetime

logger = logging.getLogger(__name__)


class AutoUpdater:
    """Service for automatic device configuration updates"""

    # Configurazione corretta per dispositivi PAI-PL
    DEVICE_MAPPING = {
        '192.168.2.1': '161+672',
        '192.168.2.2': '161+306',
        '192.168.2.3': '237+972',
        '192.168.2.4': '236+132',
        '192.168.2.5': '237+117',
        '192.168.2.6': '234+121',
        '192.168.2.7': '235+652',
        '192.168.2.10': '30+517',
        '192.168.2.15': '19+432',
        '192.168.2.20': '209+856',
        '192.168.2.25': '122+142',
        '192.168.2.30': '137+902',
        '192.168.2.50': '36+189',
        '192.168.2.100': '37+308',
    }

    @staticmethod
    def check_and_update_devices(session) -> tuple[bool, int, str]:
        """
        Controlla e aggiorna automaticamente la configurazione dei dispositivi

        Args:
            session: Database session

        Returns:
            (updated, count, message) tuple
        """
        try:
            from ..models.device import Device

            devices = session.query(Device).all()

            if not devices:
                logger.info("No devices found in database")
                return False, 0, "No devices to update"

            needs_update = []

            # Check if devices need update
            for device in devices:
                if device.ip_address in AutoUpdater.DEVICE_MAPPING:
                    expected_km = AutoUpdater.DEVICE_MAPPING[device.ip_address]

                    # Check if configuration is correct
                    if (device.name != "PAI-PL" or
                        device.device_type != "cpuB" or
                        device.location != expected_km or
                        not device.ping_enabled or
                        not device.http_enabled or
                        not device.ssh_enabled or
                        device.ssh_port != 22 or
                        device.http_port != 80):
                        needs_update.append(device)

            if not needs_update:
                logger.info("All devices are already correctly configured")
                return False, 0, "All devices up to date"

            # Apply updates
            logger.info(f"Updating {len(needs_update)} devices...")

            for device in needs_update:
                km = AutoUpdater.DEVICE_MAPPING[device.ip_address]

                device.name = "PAI-PL"
                device.device_type = "cpuB"
                device.location = km

                # Enable required checks
                device.ping_enabled = True
                device.http_enabled = True
                device.http_port = 80
                device.https_enabled = False
                device.ssh_enabled = True
                device.ssh_port = 22
                device.dns_enabled = False
                device.snmp_enabled = False

                # Enable alerts and recovery
                device.alert_enabled = True
                device.alert_on_down = True
                device.alert_on_up = True
                device.alert_on_degraded = True

                logger.info(f"Updated device {device.ip_address} -> {km}")

            session.commit()

            message = f"Auto-updated {len(needs_update)} devices to PAI-PL configuration"
            logger.info(message)

            return True, len(needs_update), message

        except Exception as e:
            logger.error(f"Auto-update failed: {e}", exc_info=True)
            session.rollback()
            return False, 0, f"Update failed: {str(e)}"

    @staticmethod
    def load_from_excel(excel_path: Path) -> Optional[Dict[str, str]]:
        """
        Load device mapping from Excel file if available

        Args:
            excel_path: Path to Excel file

        Returns:
            Mapping dict or None
        """
        try:
            import openpyxl

            if not excel_path.exists():
                logger.debug(f"Excel file not found: {excel_path}")
                return None

            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active

            mapping = {}

            # Skip header row, start from row 2
            for row in range(2, sheet.max_row + 1):
                pl = sheet.cell(row, 2).value  # Kilometrica
                ip = str(sheet.cell(row, 7).value) if sheet.cell(row, 7).value else None

                if ip and pl:
                    mapping[ip] = str(pl)

            logger.info(f"Loaded {len(mapping)} device mappings from Excel")
            return mapping if mapping else None

        except ImportError:
            logger.warning("openpyxl not installed, cannot load from Excel")
            return None
        except Exception as e:
            logger.error(f"Failed to load Excel: {e}")
            return None

    @staticmethod
    def auto_sync_on_startup(session) -> tuple[bool, str]:
        """
        Automatic synchronization on application startup

        Args:
            session: Database session

        Returns:
            (success, message) tuple
        """
        logger.info("=" * 80)
        logger.info("AUTO-SYNC: Checking device configuration...")
        logger.info("=" * 80)

        # Try to load from Excel if available
        desktop_path = Path.home() / "Desktop" / "lista PL.xlsx"
        if desktop_path.exists():
            logger.info(f"Found Excel file: {desktop_path}")
            excel_mapping = AutoUpdater.load_from_excel(desktop_path)

            if excel_mapping:
                # Update class mapping with Excel data
                AutoUpdater.DEVICE_MAPPING.update(excel_mapping)
                logger.info("Updated device mapping from Excel file")

        # Check and update devices
        updated, count, message = AutoUpdater.check_and_update_devices(session)

        if updated:
            logger.info(f"AUTO-SYNC: {message}")
            return True, message
        else:
            logger.info("AUTO-SYNC: No updates needed")
            return False, "Configuration already up to date"
